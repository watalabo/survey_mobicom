#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import argparse
import sys
import io
import re

def illegal_char_remover(data):
    ILLEGAL_CHARACTERS_RE = re.compile(
        r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]')
    """Remove ILLEGAL CHARACTER."""
    if isinstance(data, str):
        return ILLEGAL_CHARACTERS_RE.sub("", data)
    else:
        return data

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

parser = argparse.ArgumentParser(description="excel hoge hoge")
parser.add_argument("filename", help="filename hog hoge")

args = parser.parse_args()
filename = args.filename

fp = open(filename, "r", encoding="utf-8")

text = fp.read()

index = 0


wb = openpyxl.Workbook()
ws = wb.active

# print only value is set
i = 1
ret = ws.cell(row=i, column=1).value
while ret != None:
    print(ret)
    i = i + 1
    ret = ws.cell(row=i, column=1).value


row_excel = 1

index_start = text.find('<td class="tg-31dk">', index)
while index_start != -1:
    index_end = text.find('<a href', index_start)

    print(index_start)
    print(index_end)

    print(text[index_start:index_end])
    target = text[index_start:index_end]
    target = target.strip()

    print(target)
    exit(1)

    start_tmp = target.find('<a', 0)
    start_tmp = target.find('">', start_tmp)
    end_tmp = target.find('</a>', start_tmp)

    title = target[start_tmp + 2:end_tmp]
    title = title.strip()
    p = re.compile(r"<[^>]*?>")
    title = p.sub("", title)
    print(title)
    start_author = target.find('<td class="tg-c6of">')
    start_author += len('<td class="tg-c6of">')
    end_author = target.find('</td>', start_author)
    text_author = target[start_author:end_author]
    text_author = illegal_char_remover(text_author)
    print(text_author)
    ws.cell(row=row_excel, column=1).value = title
    ws.cell(row=row_excel, column=2).value = text_author
    index_start = text.find('<td class="tg-31dk">', index_end)
    row_excel = row_excel + 1


wb.save(filename + ".xlsx")    
